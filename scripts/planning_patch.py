from pathlib import Path
import csv
import datetime
import json
import math
import re
import unicodedata

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
INPUT_DIR = ROOT / "input"
DOCS_DIR = ROOT / "docs"


def norm_header(value):
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ").replace("\xa0", " ")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def norm_pdv(value):
    nums = re.findall(r"\d+", str(value or ""))
    return nums[0].zfill(5) if nums else ""


def clean(value):
    return "" if value is None else str(value).strip()


def as_float(value):
    if value in (None, ""):
        return None
    try:
        out = float(str(value).strip().replace(",", "."))
    except Exception:
        return None
    return out if math.isfinite(out) else None


def latest_file(folder, suffixes=(".xlsx", ".csv")):
    if not folder.exists():
        return None
    files = [p for p in folder.rglob("*") if p.suffix.lower() in suffixes and not p.name.startswith("~$")]
    return sorted(files, key=lambda p: p.stat().st_mtime)[-1] if files else None


def find_header(ws):
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        hs = [norm_header(ws.cell(row_idx, c).value) for c in range(1, ws.max_column + 1)]
        has_pv = any(h in {"PV", "PDV", "PV ENI", "N PV", "N PV ENI"} or h.startswith("PV ") for h in hs)
        has_place = any("CITTA" in h or "COMUNE" in h or "INDIRIZZO" in h for h in hs)
        if has_pv and has_place:
            return row_idx
    return 1


def col_map(ws, header_row):
    return {norm_header(ws.cell(header_row, c).value): c for c in range(1, ws.max_column + 1) if norm_header(ws.cell(header_row, c).value)}


def find_col(headers, *names, avoid=()):
    needles = [norm_header(x) for x in names]
    avoid_n = [norm_header(x) for x in avoid]
    for n in needles:
        if n in headers and not any(a in n for a in avoid_n):
            return headers[n]
    for h, c in headers.items():
        if any(a and a in h for a in avoid_n):
            continue
        for n in needles:
            if n and n in h:
                return c
    return None


def load_lista():
    path = latest_file(INPUT_DIR / "lista", (".xlsx",))
    out = {}
    if not path:
        return out
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        pdv = norm_pdv(row[0] if row else None)
        if not pdv:
            continue
        out[pdv] = {
            "agent": clean(row[9] if len(row) > 9 else ""),
            "region": clean(row[4] if len(row) > 4 else ""),
            "province": clean((row[5] if len(row) > 5 else "") or (row[2] if len(row) > 2 else "")),
            "city": clean((row[6] if len(row) > 6 else "") or (row[3] if len(row) > 3 else "")),
            "address": clean(row[7] if len(row) > 7 else ""),
        }
    return out


def read_table(path):
    if not path:
        return [], [], ""
    if path.suffix.lower() == ".csv":
        text = path.read_text(encoding="utf-8-sig")
        try:
            dialect = csv.Sniffer().sniff(text[:4096], delimiters=";,\t|")
        except Exception:
            dialect = csv.excel
            dialect.delimiter = ";"
        raw = list(csv.reader(text.splitlines(), dialect))
        header_idx = 0
        for i, row in enumerate(raw[:15]):
            hs = [norm_header(x) for x in row]
            if any(h in {"PV", "PDV", "PV ENI"} for h in hs):
                header_idx = i
                break
        return [norm_header(x) for x in raw[header_idx]], raw[header_idx + 1:], path.name
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["PERIMETRO"] if "PERIMETRO" in wb.sheetnames else wb[wb.sheetnames[0]]
    header_row = find_header(ws)
    headers = [norm_header(ws.cell(header_row, c).value) for c in range(1, ws.max_column + 1)]
    rows = list(ws.iter_rows(min_row=header_row + 1, values_only=True))
    return headers, rows, path.name


def row_cell(row, col):
    if not col or col - 1 >= len(row):
        return ""
    return clean(row[col - 1])


def load_grab_set():
    path = latest_file(INPUT_DIR / "grabego")
    headers, rows, _ = read_table(path)
    if not headers:
        return set()
    pdv_col = None
    for i, h in enumerate(headers, start=1):
        if h in {"PV", "PDV", "PV ENI"} or "PUNTO" in h:
            pdv_col = i
            break
    out = set()
    for row in rows:
        pdv = norm_pdv(row_cell(row, pdv_col))
        if pdv:
            out.add(pdv)
    return out


def load_anag_points():
    lista = load_lista()
    grab = load_grab_set()
    path = latest_file(INPUT_DIR / "anagrafica", (".xlsx",))
    if not path:
        return [], ""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hr = find_header(ws)
    headers = col_map(ws, hr)
    cols = {
        "pdv": find_col(headers, "PV", "PDV", "PV ENI", "N PV", "N PV ENI"),
        "region": find_col(headers, "REGIONE"),
        "province": find_col(headers, "PROVINCIA", "PROV"),
        "city": find_col(headers, "CITTA", "COMUNE"),
        "address": find_col(headers, "INDIRIZZO", "VIA"),
        "rzv": find_col(headers, "RZV"),
        "cr": find_col(headers, "CR"),
        "focal": find_col(headers, "FOCAL POINT ENI", "COORDINATORE SERVIZI", "COORD SERVIZI"),
        "lat": find_col(headers, "LATITUDINE", avoid=("VECCHIA", "NON VERIFICATA")),
        "lng": find_col(headers, "LONGITUDINE", avoid=("VECCHIA", "NON VERIFICATA")),
        "tpoint": find_col(headers, "TPOINT", "T POINT", "TELEPASS POINT"),
        "grab": find_col(headers, "GRABNGO", "GRAB NGO", "GRAB GO", "GRAB E GO"),
    }
    points, seen = [], set()
    for row in ws.iter_rows(min_row=hr + 1, values_only=True):
        pdv = norm_pdv(row_cell(row, cols["pdv"]))
        if not pdv or pdv in seen:
            continue
        seen.add(pdv)
        li = lista.get(pdv, {})
        lat = as_float(row_cell(row, cols["lat"]))
        lng = as_float(row_cell(row, cols["lng"]))
        if lat is None or lng is None:
            continue
        gval = norm_header(row_cell(row, cols["grab"]))
        tval = norm_header(row_cell(row, cols["tpoint"]))
        points.append({
            "pdv": pdv,
            "agent": li.get("agent", ""),
            "region": row_cell(row, cols["region"]) or li.get("region", ""),
            "province": row_cell(row, cols["province"]) or li.get("province", ""),
            "city": row_cell(row, cols["city"]) or li.get("city", ""),
            "address": row_cell(row, cols["address"]) or li.get("address", ""),
            "rzv": row_cell(row, cols["rzv"]),
            "cr": row_cell(row, cols["cr"]),
            "focal": row_cell(row, cols["focal"]),
            "lat": lat,
            "lng": lng,
            "is_grab": pdv in grab or gval in {"SI", "S", "YES", "TRUE", "1", "X"},
            "is_tp": tval not in {"", "NO", "N", "FALSE", "0"},
        })
    points.sort(key=lambda x: (x.get("agent") or "ZZZ", x.get("region") or "", x.get("province") or "", x.get("city") or "", x.get("pdv") or ""))
    return points, path.name


CSS = """
.planning-form{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:10px;align-items:end}.planning-form label{font-size:12px;color:var(--muted);font-weight:800}.planning-form input,.planning-form select{width:100%;padding:9px;border-radius:10px;border:1px solid var(--line);background:#fff}.plan-actions{display:flex;gap:8px;flex-wrap:wrap;margin-top:10px}.plan-table{min-width:1350px}.plan-grab{box-shadow:inset 5px 0 0 #7c3aed}.plan-tp{box-shadow:inset 5px 0 0 var(--blue)}.plan-warn{background:#fff8e1}.plan-badge{display:inline-flex;border-radius:999px;padding:4px 8px;font-size:11px;font-weight:900;background:#eef4fb;color:#123764}.plan-badge.grab{background:#efe7ff;color:#5b21b6}.plan-summary{margin-top:10px}
"""

JS = r'''
const PLAN_MONTHS=['Gennaio','Febbraio','Marzo','Aprile','Maggio','Giugno','Luglio','Agosto','Settembre','Ottobre','Novembre','Dicembre'];
const PLAN={items:[],prev:{},source:''};
function planPoints(){return (APP.planning_data&&APP.planning_data.points)||[];}
function planAgents(){return [...new Set(planPoints().map(p=>p.agent).filter(Boolean))].sort((a,b)=>a.localeCompare(b,'it'));}
function planKm(a,b){if(!a||!b||a.lat==null||b.lat==null)return 0;const R=6371,dLat=(b.lat-a.lat)*Math.PI/180,dLon=(b.lng-a.lng)*Math.PI/180,la1=a.lat*Math.PI/180,la2=b.lat*Math.PI/180;const x=Math.sin(dLat/2)**2+Math.cos(la1)*Math.cos(la2)*Math.sin(dLon/2)**2;return 2*R*Math.atan2(Math.sqrt(x),Math.sqrt(1-x));}
function planNorm(s){return String(s||'').normalize('NFD').replace(/[\u0300-\u036f]/g,'').toLowerCase().replace(/[^a-z0-9]+/g,' ').trim();}
function planDateStr(d){return d.toLocaleDateString('it-IT',{weekday:'short',day:'2-digit',month:'2-digit',year:'numeric'});}
function planIso(d){return d.toISOString().slice(0,10);}
function easterDate(y){const a=y%19,b=Math.floor(y/100),c=y%100,d=Math.floor(b/4),e=b%4,f=Math.floor((b+8)/25),g=Math.floor((b-f+1)/3),h=(19*a+b-d-g+15)%30,i=Math.floor(c/4),k=c%4,l=(32+2*e+2*i-h-k)%7,m=Math.floor((a+11*h+22*l)/451),mo=Math.floor((h+l-7*m+114)/31),da=((h+l-7*m+114)%31)+1;return new Date(y,mo-1,da);}
function planHolidays(y){const out=new Set([`${y}-01-01`,`${y}-01-06`,`${y}-04-25`,`${y}-05-01`,`${y}-06-02`,`${y}-08-15`,`${y}-11-01`,`${y}-12-08`,`${y}-12-25`,`${y}-12-26`]);const e=easterDate(y);const p=new Date(e);p.setDate(e.getDate()+1);out.add(planIso(p));return out;}
function planWorkdays(monthValue){const [y,m]=String(monthValue||'').split('-').map(Number);if(!y||!m)return[];const hol=planHolidays(y),days=[];for(let d=new Date(y,m-1,1);d.getMonth()===m-1;d.setDate(d.getDate()+1)){const w=d.getDay(),iso=planIso(d);if(w!==0&&w!==6&&!hol.has(iso))days.push(new Date(d));}return days;}
function planStartPoint(points,start){const s=planNorm(start);if(!s)return points[0]||null;let m=points.find(p=>planNorm(p.pdv)===s||planNorm(p.city)===s);if(m)return m;m=points.find(p=>planNorm(`${p.pdv} ${p.city} ${p.address}`).includes(s));return m||points[0]||null;}
function planVisitMin(p){return p.is_grab?15:45;}
function planTravelMin(km){return Math.round((km/55)*60+8);}
function planPrevAge(p,monthValue){const d=PLAN.prev[p.pdv];if(!d)return 9999;const [y,m]=String(monthValue||'').split('-').map(Number);const cur=new Date(y,m-1,1),old=new Date(d);return Math.round((cur-old)/(1000*3600*24));}
function planOrder(points,start,monthValue){let left=points.slice();const ordered=[];let cur=start||left[0];while(left.length){left.sort((a,b)=>{const da=planKm(cur,a),db=planKm(cur,b),pa=planPrevAge(a,monthValue),pb=planPrevAge(b,monthValue);const pena=pa<45?500:pa<90?120:0,penb=pb<45?500:pb<90?120:0;return (da+pena)-(db+penb);});const n=left.shift();ordered.push(n);cur=n;}return ordered;}
function planAssign(ordered,days,start){let di=0,last=start,mins=0,count=0;const out=[];for(const p of ordered){const km=planKm(last,p),add=planTravelMin(km)+planVisitMin(p);const shouldNew=count>0&&((mins+add>540&&count>=4)||(count>=6&&km>8)||count>=8);if(shouldNew&&di<days.length-1){di++;mins=0;count=0;last=start;}
 const date=days[Math.min(di,days.length-1)]||new Date();const startMin=9*60+mins+planTravelMin(planKm(last,p));const hh=String(Math.floor(startMin/60)).padStart(2,'0'),mm=String(startMin%60).padStart(2,'0');out.push({...p,date:planIso(date),dateLabel:planDateStr(date),time:`${hh}:${mm}`,travel_km:Math.round(km),visit_min:planVisitMin(p),day_load:minToText(mins+add)});mins+=add;count++;last=p;}return out;}
function minToText(mins){const h=Math.floor(mins/60),m=mins%60;return `${h}h ${String(m).padStart(2,'0')}`;}
function parsePrevText(text){const map={};String(text||'').split(/\n+/).forEach(line=>{const pdv=(line.match(/\b\d{3,6}\b/)||[])[0];if(!pdv)return;let d=null;const m1=line.match(/(\d{1,2})[\/\-.](\d{1,2})[\/\-.](20\d{2})/);const m2=line.match(/(20\d{2})[\-.](\d{1,2})[\-.](\d{1,2})/);if(m1)d=`${m1[3]}-${String(m1[2]).padStart(2,'0')}-${String(m1[1]).padStart(2,'0')}`;if(m2)d=`${m2[1]}-${String(m2[2]).padStart(2,'0')}-${String(m2[3]).padStart(2,'0')}`;if(d)map[pdv.padStart(5,'0')]=d;});return map;}
function loadPrevPlanning(file){if(!file){PLAN.prev={};return;}const r=new FileReader();r.onload=()=>{PLAN.prev=parsePrevText(r.result);alert(`Storico caricato: ${Object.keys(PLAN.prev).length} PV con data trovata`);};r.readAsText(file);}
function renderPlanningAuto(){const w=document.getElementById('planningAutoWrap');if(!w)return;const agents=planAgents(),now=new Date(),month=`${now.getFullYear()}-${String(now.getMonth()+1).padStart(2,'0')}`,data=APP.planning_data||{};w.innerHTML=`<div class="card"><div class="section-title" style="margin-bottom:8px">Planning automatico</div><div class="small-muted">Fonte: ${esc(data.source_name||'anagrafica')} · Usa solo PV con coordinate presenti. Sabato, domenica e festività nazionali sono esclusi automaticamente.</div><div class="planning-form" style="margin-top:12px"><div><label>Agente</label><select id="planAgent"><option value="">Seleziona agente</option>${agents.map(a=>`<option>${esc(a)}</option>`).join('')}</select></div><div><label>Mese</label><input id="planMonth" type="month" value="${month}"></div><div><label>Punto di partenza</label><input id="planStart" placeholder="Città o PV di partenza"></div><div><label>Grab & Go</label><select id="planGrab"><option value="no">No</option><option value="yes">Sì, includili</option></select></div><div><label>Planning precedente</label><input id="planPrev" type="file" accept=".csv,.txt,.xls"></div></div><div class="plan-actions"><button class="btn" onclick="generatePlanning()">Crea planning</button><button class="btn light" onclick="downloadPlanCsv()">Scarica CSV</button><button class="btn light" onclick="downloadPlanXls()">Scarica Excel</button></div><div class="small-muted" style="margin-top:8px">Per lo storico precedente: carica un CSV o il file Excel esportato da questa pagina; il sistema legge soprattutto PV e date.</div></div><div id="planningResult"></div>`;document.getElementById('planPrev')?.addEventListener('change',e=>loadPrevPlanning(e.target.files[0]));}
function generatePlanning(){const agent=document.getElementById('planAgent')?.value||'',month=document.getElementById('planMonth')?.value||'',includeGrab=(document.getElementById('planGrab')?.value==='yes'),startText=document.getElementById('planStart')?.value||'';if(!agent||!month){alert('Scegli agente e mese');return;}let pts=planPoints().filter(p=>p.agent===agent&&(includeGrab||!p.is_grab));const noCoords=planPoints().filter(p=>p.agent===agent&&(!p.lat||!p.lng)).length;if(!pts.length){alert('Nessun PV con coordinate per questo agente');return;}const days=planWorkdays(month);const start=planStartPoint(pts,startText);const ordered=planOrder(pts,start,month);PLAN.items=planAssign(ordered,days,start);PLAN.source=`${agent}_${month}`;renderPlanningTable(noCoords,days.length);}
function recalcPlanning(){const month=document.getElementById('planMonth')?.value||'',startText=document.getElementById('planStart')?.value||'';const days=planWorkdays(month),start=planStartPoint(PLAN.items,startText);PLAN.items=planAssign(PLAN.items,days,start);renderPlanningTable(0,days.length);}
function movePlan(i,d){const j=i+d;if(j<0||j>=PLAN.items.length)return;[PLAN.items[i],PLAN.items[j]]=[PLAN.items[j],PLAN.items[i]];recalcPlanning();}
function removePlan(i){PLAN.items.splice(i,1);recalcPlanning();}
function renderPlanningTable(noCoords,workdays){const w=document.getElementById('planningResult');if(!w)return;const days=[...new Set(PLAN.items.map(x=>x.date))].length,grab=PLAN.items.filter(x=>x.is_grab).length;const body=PLAN.items.map((p,i)=>`<tr class="${p.is_grab?'plan-grab':'plan-tp'}"><td><button class="btn light small" onclick="movePlan(${i},-1)">↑</button> <button class="btn light small" onclick="movePlan(${i},1)">↓</button> <button class="btn ghost small" onclick="removePlan(${i})">x</button></td><td>${esc(p.dateLabel)}</td><td>${esc(p.time)}</td><td><b>${esc(p.pdv)}</b></td><td>${p.is_grab?'<span class="plan-badge grab">Grab & Go</span>':'<span class="plan-badge">Telepass Point</span>'}</td><td><div class="city-cell"><div class="city-main">${esc(p.city)}</div><div class="city-address">${esc(p.address||'')}</div></div></td><td>${esc(p.province||'')}</td><td>${esc(p.region||'')}</td><td>${esc(p.focal||'')}</td><td>${esc(p.rzv||'')}</td><td>${esc(p.cr||'')}</td><td class="num">${fmtNum(p.travel_km||0)} km</td><td class="num">${fmtNum(p.visit_min||0)} min</td><td>${esc(p.day_load||'')}</td></tr>`).join('');w.innerHTML=`<div class="metric-row sost plan-summary"><div class="metric-card"><h4>PV pianificati</h4><div class="metric-big">${fmtNum(PLAN.items.length)}</div><div class="metric-sub">Grab & Go: ${fmtNum(grab)}</div></div><div class="metric-card"><h4>Giorni usati</h4><div class="metric-big">${fmtNum(days)}</div><div class="metric-sub">Lavorativi mese: ${fmtNum(workdays||0)}</div></div><div class="metric-card ${noCoords?'tone-warn':''}"><h4>Senza coordinate</h4><div class="metric-big">${fmtNum(noCoords||0)}</div><div class="metric-sub">Da sistemare in anagrafica</div></div></div><div class="card"><div class="small-muted" style="margin-bottom:8px">Puoi spostare le righe su/giù: data e orario vengono ricalcolati automaticamente.</div><div class="list-wrap"><table class="plan-table"><thead><tr><th>Modifica</th><th>Data</th><th>Ora</th><th>PV</th><th>Tipo</th><th>Città / indirizzo</th><th>Prov.</th><th>Regione</th><th>Focal Point ENI</th><th>RZV</th><th>CR</th><th>Km stimati</th><th>Visita</th><th>Carico giorno</th></tr></thead><tbody>${body}</tbody></table></div></div>`;}
function planExportRows(){return [['DATA','ORA','n° PV','TIPO','REGIONE','PROVINCIA','CITTA','INDIRIZZO','FOCAL POINT ENI','RZV','CR','KM STIMATI','DURATA VISITA','NOTE']].concat(PLAN.items.map(p=>[p.dateLabel,p.time,p.pdv,p.is_grab?'GRAB & GO':'TELEPASS POINT',p.region||'',p.province||'',p.city||'',p.address||'',p.focal||'',p.rzv||'',p.cr||'',p.travel_km||0,p.visit_min+' min',p.is_grab?'VISITA GRAB & GO':'']));}
function downloadPlanCsv(){if(!PLAN.items.length){alert('Prima crea il planning');return;}const csv=planExportRows().map(r=>r.map(v=>'"'+String(v??'').replace(/"/g,'""')+'"').join(';')).join('\n');const a=document.createElement('a');a.href=URL.createObjectURL(new Blob(['\ufeff'+csv],{type:'text/csv;charset=utf-8'}));a.download=`planning_${PLAN.source||'automatico'}.csv`;a.click();}
function downloadPlanXls(){if(!PLAN.items.length){alert('Prima crea il planning');return;}const rows=planExportRows();let html='<html><head><meta charset="UTF-8"></head><body><table border="1">'+rows.map((r,i)=>'<tr>'+r.map(v=>i?`<td>${esc(v)}</td>`:`<th>${esc(v)}</th>`).join('')+'</tr>').join('')+'</table></body></html>';const a=document.createElement('a');a.href=URL.createObjectURL(new Blob([html],{type:'application/vnd.ms-excel'}));a.download=`planning_${PLAN.source||'automatico'}.xls`;a.click();}
'''


def patch_html(html, data):
    if "APP.planning_data" in html:
        return html
    html = html.replace("const DATA = APP.rows || [];", f"APP.planning_data = {json.dumps(data, ensure_ascii=False)};\nconst DATA = APP.rows || [];", 1)
    html = html.replace(".export-note{font-size:12px;color:var(--muted);margin-top:8px;text-align:right}", ".export-note{font-size:12px;color:var(--muted);margin-top:8px;text-align:right}\n" + CSS, 1)
    if "data-page=\"grab-go\"" in html:
        html = html.replace("<button data-page=\"grab-go\" onclick=\"showPage('grab-go', this)\">Grab & Go</button>", "<button data-page=\"planning-auto\" onclick=\"showPage('planning-auto', this)\">Planning automatico</button>\n      <button data-page=\"grab-go\" onclick=\"showPage('grab-go', this)\">Grab & Go</button>", 1)
    else:
        html = html.replace("<button data-page=\"file-utili\" onclick=\"showPage('file-utili', this)\">File utili</button>", "<button data-page=\"planning-auto\" onclick=\"showPage('planning-auto', this)\">Planning automatico</button>\n      <button data-page=\"file-utili\" onclick=\"showPage('file-utili', this)\">File utili</button>", 1)
    html = html.replace("  <section id=\"page-file-utili\" class=\"page\">", "  <section id=\"page-planning-auto\" class=\"page\">\n    <div class=\"section-title\">Planning automatico</div>\n    <div id=\"planningAutoWrap\"></div>\n  </section>\n\n  <section id=\"page-file-utili\" class=\"page\">", 1)
    html = html.replace("function renderFiles(){", JS + "\nfunction renderFiles(){", 1)
    html = html.replace("if(page==='file-utili') renderFiles();", "if(page==='planning-auto') renderPlanningAuto();\n  if(page==='file-utili') renderFiles();", 1)
    html = html.replace("if(activePage()==='gare-pdv') renderGarePdv();", "if(activePage()==='gare-pdv') renderGarePdv();\n  if(activePage()==='planning-auto') renderPlanningAuto();", 1)
    html = html.replace("renderGarePdv();\nrenderGrabGo();\nrenderGare('gareAgentiWrap',APP.gare_agenti,'Nessuna gara in corso');", "renderGarePdv();\nrenderGrabGo();\nrenderPlanningAuto();\nrenderGare('gareAgentiWrap',APP.gare_agenti,'Nessuna gara in corso');", 1)
    return html


def main():
    points, source_name = load_anag_points()
    data = {
        "source_name": source_name,
        "updated_at": datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
        "points": points,
        "summary": {"points_with_coordinates": len(points), "agents": len({p.get("agent") for p in points if p.get("agent")})},
    }
    for name in ["index.html", "Telepass_ENI_sito_v6.html"]:
        path = DOCS_DIR / name
        if path.exists():
            path.write_text(patch_html(path.read_text(encoding="utf-8"), data), encoding="utf-8")
    print(f"Planning patch completata: {len(points)} PV con coordinate")


if __name__ == "__main__":
    main()
