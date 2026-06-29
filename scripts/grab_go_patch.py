from pathlib import Path
import csv
import datetime
import json
import re
import shutil
import unicodedata

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
GRABEGO_DIR = ROOT / "input" / "grabego"
LISTA_DIR = ROOT / "input" / "lista"
DOCS_DIR = ROOT / "docs"


def norm_header(value):
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ").replace("\xa0", " ")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = re.sub(r"[^A-Z0-9%]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def norm_pdv(value):
    nums = re.findall(r"\d+", str(value or ""))
    return nums[0].zfill(5) if nums else ""


def cell(row, idx):
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    return "" if val is None else str(val).strip()


def find_col(headers, *needles):
    normalized_needles = [norm_header(n) for n in needles if n]
    for i, h in enumerate(headers):
        for n in normalized_needles:
            if n and n in h:
                return i
    return None


def latest_file(folder, suffixes=(".xlsx", ".csv")):
    if not folder.exists():
        return None
    files = [p for p in folder.rglob("*") if p.suffix.lower() in suffixes and not p.name.startswith("~$")]
    if not files:
        return None
    preferred = [p for p in files if p.stem.upper().replace(" ", "").replace("_", "").replace("-", "").startswith("GRABEGO")]
    return sorted(preferred or files, key=lambda p: p.stat().st_mtime)[-1]


def load_lista_map():
    path = latest_file(LISTA_DIR, (".xlsx",))
    if not path:
        return {}
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pdv = norm_pdv(row[0] if row else "")
        if not pdv:
            continue
        # Nella lista principale l'agente è quello storico usato da Home/Classifica.
        agent = (row[9] if len(row) > 9 and row[9] is not None else "")
        out[pdv] = {"agent": str(agent).strip()}
    return out


def agent_from_email(value):
    text = str(value or "").strip()
    if "@" not in text:
        return ""
    local = re.sub(r"[._-]+", " ", text.split("@", 1)[0])
    local = re.sub(r"\s+", " ", local).strip()
    return " ".join(part.capitalize() for part in local.split())


def read_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["PERIMETRO"] if "PERIMETRO" in wb.sheetnames else wb[wb.sheetnames[0]]
    header_row = 1
    for r in range(1, min(ws.max_row, 15) + 1):
        hs = [norm_header(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if any(h in {"PV", "PDV", "PV ENI"} or "PUNTO" in h for h in hs) and any("AGENTE" in h for h in hs):
            header_row = r
            break
    headers = [norm_header(ws.cell(header_row, c).value) for c in range(1, ws.max_column + 1)]
    rows = list(ws.iter_rows(min_row=header_row + 1, values_only=True))
    return headers, rows


def read_csv(path):
    text = path.read_text(encoding="utf-8-sig")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=";,\t|")
    except Exception:
        dialect = csv.excel
        dialect.delimiter = ";"
    raw = list(csv.reader(text.splitlines(), dialect))
    header_idx = 0
    for i, candidate in enumerate(raw[:15]):
        hs = [norm_header(v) for v in candidate]
        if any(h in {"PV", "PDV", "PV ENI"} or "PUNTO" in h for h in hs) and any("AGENTE" in h for h in hs):
            header_idx = i
            break
    return [norm_header(v) for v in raw[header_idx]], raw[header_idx + 1:]


def load_grab_go():
    path = latest_file(GRABEGO_DIR)
    lista_map = load_lista_map()
    lista_pdv = set(lista_map.keys())
    if not path:
        return {
            "missing_file": True,
            "source_name": "",
            "source_path": "",
            "updated_at": "",
            "rows": [],
            "summary": {"pdv_count": 0, "telepass_point": 0, "not_telepass_point": 0, "agents": 0, "with_coords": 0},
        }

    headers, raw_rows = read_csv(path) if path.suffix.lower() == ".csv" else read_xlsx(path)
    cols = {
        "pdv": find_col(headers, "PV ENI", "PDV", "PV", "PUNTO EROGAZIONE"),
        "contract": find_col(headers, "CONTRATTO"),
        "address": find_col(headers, "INDIRIZZO"),
        "city": find_col(headers, "CITTA"),
        "region": find_col(headers, "REGIONE"),
        "agent_email": find_col(headers, "AGENTE"),
        "branch": find_col(headers, "FILIALE"),
        "stand": find_col(headers, "STAND"),
        "note": find_col(headers, "NOTE TELEPASS", "NOTE"),
        "manager_mail": find_col(headers, "MAIL GESTORE"),
        "lat": find_col(headers, "LATITUDINE", "LATITUDE", "LAT"),
        "lng": find_col(headers, "LONGITUDINE", "LONGITUDE", "LNG", "LON"),
    }
    if cols["pdv"] is None:
        cols["pdv"] = 0
    # Richiesta utente: colonna F = mail agente. Se non trovo l'header, uso F.
    if cols["agent_email"] is None:
        cols["agent_email"] = 5

    rows, seen = [], set()
    for row in raw_rows:
        pdv = norm_pdv(cell(row, cols["pdv"]))
        if not pdv or pdv in seen:
            continue
        seen.add(pdv)
        mail = cell(row, cols["agent_email"])
        agent_from_lista = lista_map.get(pdv, {}).get("agent", "")
        agent_from_mail = agent_from_email(mail)
        rows.append({
            "pdv": pdv,
            "contract": cell(row, cols["contract"]),
            "address": cell(row, cols["address"]),
            "city": cell(row, cols["city"]),
            "region": cell(row, cols["region"]),
            "agent_email": mail,
            "agent": agent_from_lista or agent_from_mail,
            "agent_raw": agent_from_mail,
            "branch": cell(row, cols["branch"]),
            "stand": cell(row, cols["stand"]),
            "note": cell(row, cols["note"]),
            "manager_mail": cell(row, cols["manager_mail"]),
            "telepass_point": pdv in lista_pdv,
            "lat": cell(row, cols["lat"]),
            "lng": cell(row, cols["lng"]),
        })
    rows.sort(key=lambda r: (r.get("agent") or "ZZZ", r.get("region") or "", r.get("city") or "", r.get("pdv") or ""))

    dest_dir = DOCS_DIR / "files" / "GRABEGO"
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest_name = "GRABEGO" + path.suffix.lower()
    shutil.copy2(path, dest_dir / dest_name)

    with_coords = 0
    for row in rows:
        try:
            float(str(row.get("lat", "")).replace(",", "."))
            float(str(row.get("lng", "")).replace(",", "."))
            with_coords += 1
        except Exception:
            pass

    return {
        "missing_file": False,
        "source_name": path.name,
        "source_path": f"files/GRABEGO/{dest_name}",
        "updated_at": datetime.datetime.fromtimestamp(path.stat().st_mtime).strftime("%d/%m/%Y %H:%M"),
        "rows": rows,
        "summary": {
            "pdv_count": len(rows),
            "telepass_point": sum(1 for r in rows if r["telepass_point"]),
            "not_telepass_point": sum(1 for r in rows if not r["telepass_point"]),
            "agents": len({r["agent"] for r in rows if r["agent"]}),
            "with_coords": with_coords,
        },
    }


CSS = """
.grab-filters{margin-bottom:10px}.grab-agent-card{margin-top:12px}.grab-table{min-width:1250px}.grab-table select{width:150px;padding:7px;border-radius:10px;border:1px solid var(--line);background:#fff}.grab-green{box-shadow:inset 5px 0 0 var(--green)}.grab-red{box-shadow:inset 5px 0 0 var(--red)}.grab-yellow{box-shadow:inset 5px 0 0 #d6a500}.grab-dot{display:inline-flex;align-items:center;justify-content:center;width:34px;height:34px;border-radius:50%;background:#0f2746;color:#fff;font-weight:900}.grab-map{margin-top:10px}.grab-map-box{height:520px;border:1px solid var(--line);border-radius:16px;overflow:hidden;background:#eef4fb;margin-top:10px}.grab-map-box .leaflet-container{height:100%;width:100%}.grab-marker{width:34px;height:34px;border-radius:50%;display:flex;align-items:center;justify-content:center;color:#fff;font-weight:900;font-size:12px;border:3px solid #fff;box-shadow:0 2px 9px rgba(0,0,0,.35)}.grab-marker.green{background:#138a36}.grab-marker.red{background:#c62828}.grab-marker.yellow{background:#d6a500;color:#10243e}.grab-legend{display:flex;gap:8px;flex-wrap:wrap;margin-top:8px}.grab-legend span{display:inline-flex;align-items:center;gap:6px;background:#f1f5fb;border:1px solid var(--line);border-radius:999px;padding:6px 10px;font-size:12px;font-weight:800}.grab-legend i{width:12px;height:12px;border-radius:50%;display:inline-block}.grab-popup b{font-size:14px}.grab-popup .muted{font-size:12px}.grab-no-coords{margin-top:8px;border:1px solid #f2d68c;background:#fff8e6;color:#8f5c00;padding:10px 12px;border-radius:12px;font-size:12px}
"""

JS = r'''
const GRAB_GO_MONTHS=['Gennaio','Febbraio','Marzo','Aprile','Maggio','Giugno','Luglio','Agosto','Settembre','Ottobre','Novembre','Dicembre'];
function grabRows(){return (APP.grab_go&&APP.grab_go.rows)||[];}
function grabState(){try{return JSON.parse(localStorage.getItem('grabGoVisits')||'{}')||{};}catch(e){return{};}}
function saveGrabState(s){localStorage.setItem('grabGoVisits',JSON.stringify(s||{}));}
function grabNameParts(s){return String(s||'').normalize('NFD').replace(/[\u0300-\u036f]/g,'').toLowerCase().replace(/[^a-z0-9 ]+/g,' ').replace(/\s+/g,' ').trim().split(' ').filter(Boolean);}
function grabSameTokens(a,b){if(a.length!==b.length)return false;return [...a].sort().join('|')===[...b].sort().join('|');}
function grabInitialCompatible(raw,site){if(!raw.length||!site.length)return false;if(raw[0]!==site[0])return false;for(let i=1;i<raw.length;i++){const r=raw[i],s=site[i];if(!s)return false;if(r===s)continue;if(r.length===1&&s.startsWith(r))continue;return false;}return true;}
function grabCanonicalAgent(raw){const clean=String(raw||'').trim();if(!clean)return '';const rawParts=grabNameParts(clean),rawNorm=rawParts.join(' ');const site=[...new Set((DATA||[]).map(r=>r.agent).filter(Boolean))];for(const a of site){const p=grabNameParts(a);if(p.join(' ')===rawNorm)return a;}for(const a of site){const p=grabNameParts(a);if(grabSameTokens(rawParts,p))return a;}for(const a of site){const p=grabNameParts(a);if(grabInitialCompatible(rawParts,p))return a;if(grabSameTokens(rawParts.slice().reverse(),p))return a;if(grabInitialCompatible(rawParts.slice().reverse(),p))return a;}return clean;}
function grabAgent(r){if(!r._agentCanonical)r._agentCanonical=grabCanonicalAgent(r.agent||r.agent_raw||'');return r._agentCanonical;}
function grabAgentOptions(){return [...new Set(grabRows().map(grabAgent).filter(Boolean))].sort((a,b)=>String(a).localeCompare(String(b),'it'));}
function grabInitials(a){return String(a||'?').split(/\s+/).filter(Boolean).slice(0,2).map(x=>x[0]).join('').toUpperCase()||'?';}
function setGrabVisit(pdv,m){const s=grabState();if(!m){delete s[pdv];}else{s[pdv]={month:Number(m),year:new Date().getFullYear(),saved_at:new Date().toISOString()};}saveGrabState(s);renderGrabGo();}
function grabVisit(pdv){const x=grabState()[pdv];if(!x||!x.month)return{cls:'grab-red',color:'red',label:'Non visitato',rank:2};const now=new Date(),m=Number(x.month),y=Number(x.year||now.getFullYear()),diff=(now.getFullYear()-y)*12+((now.getMonth()+1)-m),label=`Visitato: ${GRAB_GO_MONTHS[m-1]||m} ${y}`;if(diff>4)return{cls:'grab-yellow',color:'yellow',label:'Visitato più di 4 mesi fa · '+label,rank:1};return{cls:'grab-green',color:'green',label,rank:0};}
function grabCoord(v){const n=Number(String(v??'').replace(',','.').trim());return Number.isFinite(n)?n:null;}
function grabHasCoord(r){const lat=grabCoord(r.lat),lng=grabCoord(r.lng);return lat!==null&&lng!==null&&lat>=-90&&lat<=90&&lng>=-180&&lng<=180;}
function grabQuery(r){return `${r.address||''} ${r.city||''} ${r.region||''}`.trim()||r.pdv;}
function openGrabMap(pdv){const marker=window.grabMarkers&&window.grabMarkers[pdv];if(marker&&window.grabLeafletMap){window.grabLeafletMap.setView(marker.getLatLng(),14,{animate:true});marker.openPopup();document.getElementById('grabMapBox')?.scrollIntoView({behavior:'smooth',block:'center'});}}
function grabFiltered(){const q=(document.getElementById('searchText')?.value||'').toLowerCase().trim(),agent=document.getElementById('grabAgentFilter')?.value||'',visit=document.getElementById('grabVisitFilter')?.value||'',tp=document.getElementById('grabTpFilter')?.value||'',stand=document.getElementById('grabStandFilter')?.value||'';return grabRows().filter(r=>{const ag=grabAgent(r);if(agent&&ag!==agent)return false;if(tp==='yes'&&!r.telepass_point)return false;if(tp==='no'&&r.telepass_point)return false;if(stand&&String(r.stand||'')!==stand)return false;const vi=grabVisit(r.pdv);if(visit==='not'&&vi.rank!==2)return false;if(visit==='recent'&&vi.rank!==0)return false;if(visit==='old'&&vi.rank!==1)return false;if(q){const hay=`${r.pdv} ${r.city} ${r.address} ${r.region} ${ag} ${r.agent_email} ${r.contract} ${r.stand} ${r.note} ${r.manager_mail}`.toLowerCase();if(!hay.includes(q))return false;}return true;});}
function grabIcon(r){const vi=grabVisit(r.pdv);return L.divIcon({className:'',html:`<div class="grab-marker ${vi.color}">${esc(grabInitials(grabAgent(r)))}</div>`,iconSize:[34,34],iconAnchor:[17,17],popupAnchor:[0,-17]});}
function initGrabAllMap(rows){const box=document.getElementById('grabAllMap');if(!box)return;if(!window.L){box.innerHTML='<div class="empty">Mappa non caricata. Controlla la connessione e ricarica.</div>';return;}if(window.grabLeafletMap){window.grabLeafletMap.remove();window.grabLeafletMap=null;}window.grabMarkers={};const mapped=rows.filter(grabHasCoord);const map=L.map('grabAllMap',{scrollWheelZoom:false});window.grabLeafletMap=map;L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png',{maxZoom:19,attribution:'&copy; OpenStreetMap'}).addTo(map);if(!mapped.length){map.setView([42.5,12.5],5);return;}const bounds=[];mapped.forEach(r=>{const lat=grabCoord(r.lat),lng=grabCoord(r.lng),vi=grabVisit(r.pdv),q=grabQuery(r),ag=grabAgent(r);const popup=`<div class="grab-popup"><b>${esc(r.pdv)} · ${esc(r.city||'')}</b><div class="muted">${esc(r.address||'')}</div><div style="margin-top:6px"><b>Agente:</b> ${esc(ag||'')}</div><div><b>Stato visita:</b> ${esc(vi.label)}</div><div><b>Telepass:</b> ${r.telepass_point?'Già Telepass Point':'Non TP'}</div><div style="margin-top:8px"><a target="_blank" href="https://www.google.com/maps/search/?api=1&query=${encodeURIComponent(q)}">Apri su Google Maps</a></div></div>`;const marker=L.marker([lat,lng],{icon:grabIcon(r)}).addTo(map).bindPopup(popup);window.grabMarkers[r.pdv]=marker;bounds.push([lat,lng]);});if(bounds.length===1){map.setView(bounds[0],13);}else{map.fitBounds(bounds,{padding:[30,30]});}}
function renderGrabMap(rows){const mapped=rows.filter(grabHasCoord).length,missing=rows.length-mapped;return `<div class="card grab-map"><div style="display:flex;justify-content:space-between;gap:10px;align-items:flex-start;flex-wrap:wrap"><div><h3 style="text-align:left;margin-bottom:4px">Mappa generale Grab & Go</h3><div class="small-muted">Mostra tutti i PDV filtrati insieme. I cerchi hanno le iniziali dell'agente e il colore dello stato visita.</div><div class="small-muted" style="margin-top:4px">Marker visibili: <b>${fmtNum(mapped)}</b> / ${fmtNum(rows.length)}</div></div><div class="grab-legend"><span><i style="background:#138a36"></i>Visitato recente</span><span><i style="background:#c62828"></i>Non visitato</span><span><i style="background:#d6a500"></i>Oltre 4 mesi</span></div></div><div id="grabMapBox" class="grab-map-box"><div id="grabAllMap" style="height:100%;width:100%"></div></div>${missing?`<div class="grab-no-coords">${fmtNum(missing)} PDV non hanno coordinate e quindi non compaiono sulla mappa. Per vederli tutti, aggiungi nel file colonne <b>LATITUDINE</b> e <b>LONGITUDINE</b>.</div>`:''}</div>`;}
function renderGrabGo(){const w=document.getElementById('grabGoWrap');if(!w)return;const rep=APP.grab_go||{rows:[],summary:{}};if(rep.missing_file){w.innerHTML='<div class="empty">Nessun file GRABEGO.xlsx trovato in input/grabego.</div>';return;}const ca=document.getElementById('grabAgentFilter')?.value||'',cv=document.getElementById('grabVisitFilter')?.value||'',ct=document.getElementById('grabTpFilter')?.value||'',cs=document.getElementById('grabStandFilter')?.value||'',all=grabRows(),agents=grabAgentOptions(),stands=[...new Set(all.map(r=>r.stand).filter(Boolean))].sort((a,b)=>String(a).localeCompare(String(b),'it')),rows=grabFiltered(),tp=rows.filter(r=>r.telepass_point).length,state=rows.reduce((a,r)=>{const k=grabVisit(r.pdv).rank;a[k]=(a[k]||0)+1;return a;},{}),groups=new Map();rows.forEach(r=>{const k=grabAgent(r)||'Senza agente';if(!groups.has(k))groups.set(k,[]);groups.get(k).push(r);});const opts=(vals,ph,sel)=>`<option value="">${ph}</option>`+vals.map(v=>`<option value="${esc(v)}" ${String(v)===String(sel)?'selected':''}>${esc(v)}</option>`).join('');const monthOpts=sel=>'<option value="">Non visitato</option>'+GRAB_GO_MONTHS.map((m,i)=>`<option value="${i+1}" ${Number(sel)===(i+1)?'selected':''}>${m}</option>`).join('');let html=`<div class="card grab-filters"><div class="filter-main"><select id="grabAgentFilter" onchange="renderGrabGo()">${opts(agents,'Tutti gli agenti Grab & Go',ca)}</select><select id="grabVisitFilter" onchange="renderGrabGo()"><option value="">Tutti gli stati visita</option><option value="not" ${cv==='not'?'selected':''}>Non visitati</option><option value="recent" ${cv==='recent'?'selected':''}>Visitati ultimi 4 mesi</option><option value="old" ${cv==='old'?'selected':''}>Visitati da più di 4 mesi</option></select><select id="grabTpFilter" onchange="renderGrabGo()"><option value="">Tutti TP / non TP</option><option value="yes" ${ct==='yes'?'selected':''}>Già Telepass Point</option><option value="no" ${ct==='no'?'selected':''}>Non Telepass Point</option></select><select id="grabStandFilter" onchange="renderGrabGo()">${opts(stands,'Tutti gli stand',cs)}</select></div><div class="filter-actions"><a class="btn light" href="${esc(rep.source_path||'#')}" download>Scarica GRABEGO</a></div><div class="small-muted" style="margin-top:8px">Fonte: ${esc(rep.source_name||'GRABEGO.xlsx')} · Aggiornato: ${esc(rep.updated_at||'')} · Filtro agenti separato da Home/Classifica. La ricerca libera in alto resta attiva anche qui.</div></div><div class="metric-row sost" style="margin-top:10px"><div class="metric-card"><h4>PDV Grab & Go filtrati</h4><div class="metric-big">${fmtNum(rows.length)}</div><div class="metric-sub">Totali file: ${fmtNum(all.length)}</div></div><div class="metric-card tone-good"><h4>Già Telepass Point</h4><div class="metric-big">${fmtNum(tp)}</div></div><div class="metric-card tone-bad"><h4>Non visitati</h4><div class="metric-big">${fmtNum(state[2]||0)}</div></div></div>${renderGrabMap(rows)}`;html += [...groups.entries()].sort((a,b)=>a[0].localeCompare(b[0],'it')).map(([agent,items])=>`<div class="card grab-agent-card"><div style="display:flex;justify-content:space-between;gap:10px;align-items:center;flex-wrap:wrap"><div class="section-title" style="margin:0">${esc(agent)}</div><div class="small-muted">${fmtNum(items.length)} PDV · ${fmtNum(items.filter(x=>x.telepass_point).length)} già TP</div></div><div class="list-wrap" style="margin-top:10px"><table class="grab-table"><thead><tr><th></th><th>PDV</th><th>Città / indirizzo</th><th>Telepass Point</th><th>Contratto</th><th>Stand</th><th>Filiale</th><th>Note</th><th>Visita</th><th>Mappa</th></tr></thead><tbody>${items.map(r=>{const ag=grabAgent(r),vi=grabVisit(r.pdv),st=grabState()[r.pdv]||{},has=grabHasCoord(r),q=grabQuery(r);return `<tr class="${vi.cls}"><td><span class="grab-dot">${esc(grabInitials(ag))}</span></td><td><b>${esc(r.pdv)}</b></td><td><div class="city-cell"><div class="city-main">${esc(r.city)}</div>${r.address?'<div class="city-address">'+esc(r.address)+'</div>':''}<div class="small-muted">${esc(ag||'')}</div></div></td><td>${r.telepass_point?'<span class="badge bene">Già Telepass Point</span>':'<span class="badge male">Non TP</span>'}</td><td>${esc(r.contract||'')}</td><td>${esc(r.stand||'')}</td><td>${esc(r.branch||'')}</td><td>${esc(r.note||'')}</td><td><select onchange="setGrabVisit('${r.pdv}',this.value)">${monthOpts(st.month)}</select><div class="small-muted">${esc(vi.label)}</div></td><td>${has?`<button class="btn light small" onclick="openGrabMap('${r.pdv}')">Mappa</button>`:''}<a class="btn ghost small" target="_blank" href="https://www.google.com/maps/search/?api=1&query=${encodeURIComponent(q)}">Google</a></td></tr>`;}).join('')}</tbody></table></div></div>`).join('')||'<div class="empty">Nessun PDV Grab & Go con i filtri scelti.</div>';w.innerHTML=html;setTimeout(()=>initGrabAllMap(rows),80);}
'''


def patch_html(html, data):
    if "APP.grab_go" in html:
        return html
    leaflet_css = '<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">'
    leaflet_js = '<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>'
    html = html.replace("</head>", leaflet_css + "\n</head>", 1)
    html = html.replace("</body>", leaflet_js + "\n</body>", 1)
    html = html.replace("const DATA = APP.rows || [];", f"APP.grab_go = {json.dumps(data, ensure_ascii=False)};\nconst DATA = APP.rows || [];", 1)
    html = html.replace(".export-note{font-size:12px;color:var(--muted);margin-top:8px;text-align:right}", ".export-note{font-size:12px;color:var(--muted);margin-top:8px;text-align:right}\n" + CSS, 1)
    html = html.replace("<button data-page=\"file-utili\" onclick=\"showPage('file-utili', this)\">File utili</button>", "<button data-page=\"grab-go\" onclick=\"showPage('grab-go', this)\">Grab & Go</button>\n      <button data-page=\"file-utili\" onclick=\"showPage('file-utili', this)\">File utili</button>", 1)
    html = html.replace("  <section id=\"page-file-utili\" class=\"page\">", "  <section id=\"page-grab-go\" class=\"page\">\n    <div class=\"section-title\">Grab & Go</div>\n    <div id=\"grabGoWrap\"></div>\n  </section>\n\n  <section id=\"page-file-utili\" class=\"page\">", 1)
    # Non modifichiamo più il filtro agenti globale: Home/Classifica restano con gli agenti originali.
    html = html.replace("function renderFiles(){", JS + "\nfunction renderFiles(){", 1)
    html = html.replace("if(page==='file-utili') renderFiles();", "if(page==='grab-go') renderGrabGo();\n  if(page==='file-utili') renderFiles();", 1)
    html = html.replace("if(activePage()==='gare-pdv') renderGarePdv();", "if(activePage()==='gare-pdv') renderGarePdv();\n  if(activePage()==='grab-go') renderGrabGo();", 1)
    html = html.replace("renderGarePdv();\nrenderGare('gareAgentiWrap',APP.gare_agenti,'Nessuna gara in corso');", "renderGarePdv();\nrenderGrabGo();\nrenderGare('gareAgentiWrap',APP.gare_agenti,'Nessuna gara in corso');", 1)
    return html


def main():
    data = load_grab_go()
    for name in ["index.html", "Telepass_ENI_sito_v6.html"]:
        path = DOCS_DIR / name
        if path.exists():
            path.write_text(patch_html(path.read_text(encoding="utf-8"), data), encoding="utf-8")
    print(f"Grab & Go patch completata: {len(data.get('rows', []))} PDV, coordinate: {data.get('summary', {}).get('with_coords', 0)}")


if __name__ == "__main__":
    main()
