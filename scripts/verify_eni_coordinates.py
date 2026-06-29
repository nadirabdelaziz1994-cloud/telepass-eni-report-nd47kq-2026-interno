from __future__ import annotations

import argparse
import datetime as dt
import json
import math
import os
import re
import sys
import time
import unicodedata
from pathlib import Path
from typing import Any
from urllib.parse import urlencode
from urllib.request import Request, urlopen

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
INPUT_DIR = ROOT / "input" / "anagrafica"
OUTPUT_DIR = ROOT / "output" / "anagrafica"
CACHE_PATH = OUTPUT_DIR / "coordinate_google_cache.json"


def norm_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\r", " ").replace("\n", " ").replace("\xa0", " ")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^A-Za-z0-9]+", " ", text).upper()
    return re.sub(r"\s+", " ", text).strip()


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def parse_float(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        out = float(text)
    except Exception:
        return None
    if math.isfinite(out):
        return out
    return None


def haversine_km(a_lat: float, a_lng: float, b_lat: float, b_lng: float) -> float:
    r = 6371.0
    p1, p2 = math.radians(a_lat), math.radians(b_lat)
    dp = math.radians(b_lat - a_lat)
    dl = math.radians(b_lng - a_lng)
    x = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.atan2(math.sqrt(x), math.sqrt(1 - x))


def find_latest_xlsx(folder: Path) -> Path:
    files = [p for p in folder.rglob("*.xlsx") if not p.name.startswith("~$")]
    if not files:
        raise FileNotFoundError(f"Nessun file .xlsx trovato in {folder}")
    return max(files, key=lambda p: p.stat().st_mtime)


def detect_header(ws) -> int:
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        headers = [norm_text(ws.cell(row_idx, c).value) for c in range(1, ws.max_column + 1)]
        has_pv = any(h in {"PV", "PDV", "PV ENI", "N PV", "N PV ENI"} or h.startswith("PV ") for h in headers)
        has_city = any("CITTA" in h or "COMUNE" in h for h in headers)
        has_addr = any("INDIRIZZO" in h or "VIA" == h for h in headers)
        if has_pv and (has_city or has_addr):
            return row_idx
    return 1


def header_map(ws, header_row: int) -> dict[str, int]:
    return {norm_text(ws.cell(header_row, c).value): c for c in range(1, ws.max_column + 1) if norm_text(ws.cell(header_row, c).value)}


def find_col(headers: dict[str, int], *needles: str) -> int | None:
    ns = [norm_text(n) for n in needles]
    for n in ns:
        if n in headers:
            return headers[n]
    for key, col in headers.items():
        for n in ns:
            if n and n in key:
                return col
    return None


def ensure_col(ws, headers: dict[str, int], header_row: int, name: str) -> int:
    n = norm_text(name)
    if n in headers:
        return headers[n]
    col = ws.max_column + 1
    ws.cell(header_row, col).value = name
    headers[n] = col
    return col


def copy_old_coords_if_needed(ws, header_row: int, lat_col: int | None, lng_col: int | None, headers: dict[str, int]) -> tuple[int, int]:
    old_lat_col = ensure_col(ws, headers, header_row, "Latitudine vecchia NON VERIFICATA")
    old_lng_col = ensure_col(ws, headers, header_row, "Longitudine vecchia NON VERIFICATA")
    if lat_col:
        for r in range(header_row + 1, ws.max_row + 1):
            if ws.cell(r, old_lat_col).value in (None, ""):
                ws.cell(r, old_lat_col).value = ws.cell(r, lat_col).value
    if lng_col:
        for r in range(header_row + 1, ws.max_row + 1):
            if ws.cell(r, old_lng_col).value in (None, ""):
                ws.cell(r, old_lng_col).value = ws.cell(r, lng_col).value
    return old_lat_col, old_lng_col


def http_json(url: str, params: dict[str, str], timeout: int = 25) -> dict[str, Any]:
    full = url + "?" + urlencode(params)
    req = Request(full, headers={"User-Agent": "telepass-eni-coordinate-check/1.0"})
    with urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8"))


def load_cache() -> dict[str, Any]:
    if CACHE_PATH.exists():
        try:
            return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_cache(cache: dict[str, Any]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def score_place(place: dict[str, Any], city: str, address: str) -> tuple[int, list[str]]:
    name = clean(place.get("name"))
    formatted = clean(place.get("formatted_address"))
    types = place.get("types") or []
    text = norm_text(name + " " + formatted)
    city_n = norm_text(city)
    addr_n = norm_text(address)
    score = 0
    notes: list[str] = []

    if any(x in text for x in ["ENI", "ENILIVE", "AGIP"]):
        score += 40
        notes.append("brand trovato")
    if "gas_station" in types:
        score += 30
        notes.append("tipo gas_station")
    if city_n and city_n in text:
        score += 20
        notes.append("città trovata")
    # Non pretendiamo match perfetto sull'indirizzo perché SS/SP/KM/SNC sono scritti spesso in modi diversi.
    addr_tokens = [t for t in addr_n.split() if len(t) >= 3 and not t.isdigit()]
    if addr_tokens:
        hits = sum(1 for t in addr_tokens if t in text)
        if hits:
            score += min(20, hits * 5)
            notes.append(f"indirizzo parziale {hits}/{len(addr_tokens)}")
    return score, notes


def google_places_search(api_key: str, query: str, city: str, address: str, language: str = "it") -> dict[str, Any]:
    data = http_json(
        "https://maps.googleapis.com/maps/api/place/textsearch/json",
        {"query": query, "key": api_key, "language": language, "region": "it"},
    )
    status = data.get("status", "")
    if status not in {"OK", "ZERO_RESULTS"}:
        raise RuntimeError(f"Google Places status {status}: {data.get('error_message', '')}")
    results = data.get("results") or []
    if not results:
        return {"found": False, "status": "ZERO_RESULTS", "query": query}

    scored = []
    for place in results[:8]:
        score, notes = score_place(place, city, address)
        scored.append((score, notes, place))
    scored.sort(key=lambda x: x[0], reverse=True)
    score, notes, place = scored[0]
    loc = (((place.get("geometry") or {}).get("location")) or {})
    lat, lng = parse_float(loc.get("lat")), parse_float(loc.get("lng"))
    if lat is None or lng is None:
        return {"found": False, "status": "NO_GEOMETRY", "query": query}
    return {
        "found": True,
        "query": query,
        "score": score,
        "notes": notes,
        "name": clean(place.get("name")),
        "formatted_address": clean(place.get("formatted_address")),
        "place_id": clean(place.get("place_id")),
        "lat": lat,
        "lng": lng,
        "types": place.get("types") or [],
    }


def google_geocode_fallback(api_key: str, query: str, language: str = "it") -> dict[str, Any]:
    data = http_json(
        "https://maps.googleapis.com/maps/api/geocode/json",
        {"address": query, "key": api_key, "language": language, "region": "it"},
    )
    status = data.get("status", "")
    if status not in {"OK", "ZERO_RESULTS"}:
        raise RuntimeError(f"Google Geocode status {status}: {data.get('error_message', '')}")
    results = data.get("results") or []
    if not results:
        return {"found": False, "status": "ZERO_RESULTS", "query": query}
    item = results[0]
    loc = (((item.get("geometry") or {}).get("location")) or {})
    lat, lng = parse_float(loc.get("lat")), parse_float(loc.get("lng"))
    if lat is None or lng is None:
        return {"found": False, "status": "NO_GEOMETRY", "query": query}
    return {
        "found": True,
        "query": query,
        "score": 35,
        "notes": ["fallback geocode"],
        "name": "",
        "formatted_address": clean(item.get("formatted_address")),
        "place_id": clean(item.get("place_id")),
        "lat": lat,
        "lng": lng,
        "types": item.get("types") or [],
    }


def best_google_result(api_key: str, row_key: str, city: str, province: str, address: str, pv: str, cache: dict[str, Any], delay: float) -> dict[str, Any]:
    if row_key in cache:
        return cache[row_key]

    parts_addr = " ".join(x for x in [address, city, province, "Italia"] if x)
    queries = [
        f"Eni station {parts_addr}",
        f"Enilive {parts_addr}",
        f"Eni {parts_addr}",
    ]
    if pv:
        queries.append(f"Eni PV {pv} {city} {province} Italia")

    best: dict[str, Any] | None = None
    errors: list[str] = []
    for q in queries:
        try:
            result = google_places_search(api_key, q, city, address)
            result["method"] = "Google Places Text Search"
            if result.get("found"):
                if best is None or int(result.get("score", 0)) > int(best.get("score", 0)):
                    best = result
                if int(result.get("score", 0)) >= 65:
                    break
            else:
                best = best or result
        except Exception as exc:
            errors.append(str(exc))
        time.sleep(delay)

    if not best or not best.get("found"):
        try:
            result = google_geocode_fallback(api_key, parts_addr)
            result["method"] = "Google Geocoding fallback"
            best = result
        except Exception as exc:
            errors.append(str(exc))

    if not best:
        best = {"found": False, "status": "ERROR", "errors": errors, "query": queries[0] if queries else ""}
    if errors:
        best["errors"] = errors
    cache[row_key] = best
    save_cache(cache)
    return best


def main() -> int:
    parser = argparse.ArgumentParser(description="Verifica coordinate stazioni Eni/Enilive con Google Places.")
    parser.add_argument("--source", default="", help="Percorso Excel sorgente. Se vuoto usa l'xlsx più recente in input/anagrafica.")
    parser.add_argument("--output", default="", help="Percorso Excel output. Se vuoto scrive in output/anagrafica.")
    parser.add_argument("--max-rows", type=int, default=0, help="Limite righe per test. 0 = tutte.")
    parser.add_argument("--delay", type=float, default=0.12, help="Pausa tra chiamate Google, in secondi.")
    parser.add_argument("--min-score", type=int, default=50, help="Score minimo per accettare risultato Google Places.")
    parser.add_argument("--clear-unverified", action="store_true", help="Svuota lat/lng quando non vengono verificate.")
    args = parser.parse_args()

    api_key = os.environ.get("GOOGLE_MAPS_API_KEY", "").strip()
    if not api_key:
        print("ERRORE: manca variabile GOOGLE_MAPS_API_KEY.", file=sys.stderr)
        return 2

    source = Path(args.source) if args.source else find_latest_xlsx(INPUT_DIR)
    if not source.is_absolute():
        source = ROOT / source
    if not source.exists():
        raise FileNotFoundError(source)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if args.output:
        output = Path(args.output)
        if not output.is_absolute():
            output = ROOT / output
    else:
        stamp = dt.datetime.now().strftime("%Y%m%d_%H%M")
        output = OUTPUT_DIR / f"Anagrafica_coordinate_verificate_{stamp}.xlsx"

    wb = openpyxl.load_workbook(source)
    ws = wb[wb.sheetnames[0]]
    header_row = detect_header(ws)
    headers = header_map(ws, header_row)

    pv_col = find_col(headers, "PV", "PDV", "PV ENI", "N PV", "N PV ENI")
    address_col = find_col(headers, "INDIRIZZO", "VIA")
    city_col = find_col(headers, "CITTA", "COMUNE")
    province_col = find_col(headers, "PROVINCIA", "PROV")
    lat_col = find_col(headers, "LATITUDINE")
    lng_col = find_col(headers, "LONGITUDINE")

    if not pv_col:
        raise RuntimeError("Colonna PV/PDV non trovata.")
    if not address_col and not city_col:
        raise RuntimeError("Colonne indirizzo/città non trovate.")

    old_lat_col, old_lng_col = copy_old_coords_if_needed(ws, header_row, lat_col, lng_col, headers)
    lat_col = ensure_col(ws, headers, header_row, "Latitudine")
    lng_col = ensure_col(ws, headers, header_row, "Longitudine")
    esito_col = ensure_col(ws, headers, header_row, "Esito verifica coordinate")
    fonte_col = ensure_col(ws, headers, header_row, "Fonte coordinate")
    place_col = ensure_col(ws, headers, header_row, "Google Place ID")
    google_addr_col = ensure_col(ws, headers, header_row, "Indirizzo Google")
    dist_col = ensure_col(ws, headers, header_row, "Distanza da coordinate vecchie km")
    note_col = ensure_col(ws, headers, header_row, "Note coordinate")

    cache = load_cache()
    processed = ok = corrected = unverified = 0

    for r in range(header_row + 1, ws.max_row + 1):
        if args.max_rows and processed >= args.max_rows:
            break
        pv = clean(ws.cell(r, pv_col).value)
        address = clean(ws.cell(r, address_col).value) if address_col else ""
        city = clean(ws.cell(r, city_col).value) if city_col else ""
        province = clean(ws.cell(r, province_col).value) if province_col else ""
        if not pv and not address and not city:
            continue

        old_lat = parse_float(ws.cell(r, old_lat_col).value)
        old_lng = parse_float(ws.cell(r, old_lng_col).value)
        row_key = norm_text("|".join([pv, address, city, province]))
        result = best_google_result(api_key, row_key, city, province, address, pv, cache, args.delay)
        processed += 1

        accepted = False
        distance = None
        esito = "DA VERIFICARE"
        notes = []
        if result.get("found"):
            score = int(result.get("score", 0) or 0)
            lat, lng = parse_float(result.get("lat")), parse_float(result.get("lng"))
            if lat is not None and lng is not None:
                if old_lat is not None and old_lng is not None:
                    distance = haversine_km(old_lat, old_lng, lat, lng)
                if score >= args.min_score:
                    accepted = True
                    if distance is None:
                        esito = "VERIFICATA DA GOOGLE"
                        ok += 1
                    elif distance <= 0.35:
                        esito = "OK - coordinate vecchie confermate"
                        ok += 1
                    else:
                        esito = "CORRETTA DA GOOGLE"
                        corrected += 1
                else:
                    notes.append(f"score basso {score}")
        else:
            notes.append(str(result.get("status", "non trovata")))

        if accepted:
            ws.cell(r, lat_col).value = result.get("lat")
            ws.cell(r, lng_col).value = result.get("lng")
            ws.cell(r, fonte_col).value = result.get("method", "Google")
        else:
            unverified += 1
            if args.clear_unverified:
                ws.cell(r, lat_col).value = None
                ws.cell(r, lng_col).value = None
            ws.cell(r, fonte_col).value = "NON VERIFICATA"

        ws.cell(r, esito_col).value = esito
        ws.cell(r, place_col).value = result.get("place_id", "")
        ws.cell(r, google_addr_col).value = result.get("formatted_address", "")
        ws.cell(r, dist_col).value = round(distance, 3) if distance is not None else ""
        combined_notes = []
        if result.get("name"):
            combined_notes.append(f"Nome Google: {result.get('name')}")
        combined_notes.extend(result.get("notes") or [])
        combined_notes.extend(notes)
        if result.get("errors"):
            combined_notes.extend(result.get("errors") or [])
        ws.cell(r, note_col).value = " | ".join(str(x) for x in combined_notes if x)

        if processed % 25 == 0:
            print(f"Processati {processed} PV...")

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    summary = {
        "source": str(source.relative_to(ROOT) if source.is_relative_to(ROOT) else source),
        "output": str(output.relative_to(ROOT) if output.is_relative_to(ROOT) else output),
        "processed": processed,
        "ok_or_verified": ok,
        "corrected": corrected,
        "unverified": unverified,
        "created_at": dt.datetime.now().isoformat(timespec="seconds"),
    }
    (output.with_suffix(".json")).write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
