from pathlib import Path
import csv
import json
import math
import re
import unicodedata

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
INPUT_DIR = ROOT / "input"
DOCS_DIR = ROOT / "docs"


def norm_header(value):
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ").replace("\xa0", " ")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def clean(value):
    return "" if value is None else str(value).strip()


def norm_pdv(value):
    m = re.findall(r"\d+", str(value or ""))
    return m[0].zfill(5) if m else ""


def as_float(value):
    if value in (None, ""):
        return None
    try:
        n = float(str(value).strip().replace(",", "."))
    except Exception:
        return None
    return n if math.isfinite(n) else None


def latest_file(folder, suffixes=(".xlsx", ".csv")):
    if not folder.exists():
        return None
    files = [p for p in folder.rglob("*") if p.suffix.lower() in suffixes and not p.name.startswith("~$")]
    return sorted(files, key=lambda p: p.stat().st_mtime)[-1] if files else None


def safe_dim(value, default):
    try:
        return int(value) if value else default
    except Exception:
        return default


def find_header(ws):
    max_row = safe_dim(ws.max_row, 20)
    max_col = safe_dim(ws.max_column, 80)
    for row_idx in range(1, min(max_row, 25) + 1):
        hs = [norm_header(ws.cell(row_idx, c).value) for c in range(1, max_col + 1)]
        has_pv = any(h in {"PV", "PDV", "PV ENI", "N PV", "N PV ENI", "PUNTO EROGAZIONE"} or h.startswith("PV ") for h in hs)
        has_place = any("CITTA" in h or "COMUNE" in h or "INDIRIZZO" in h or "VIA" == h for h in hs)
        if has_pv and has_place:
            return row_idx
    return 1


def headers_for_ws(ws, header_row):
    max_col = safe_dim(ws.max_column, 80)
    return {norm_header(ws.cell(header_row, c).value): c for c in range(1, max_col + 1) if norm_header(ws.cell(header_row, c).value)}


def find_col(headers, *needles, avoid=()):
    ns = [norm_header(x) for x in needles]
    av = [norm_header(x) for x in avoid]
    for n in ns:
        if n in headers and not any(a and a in n for a in av):
            return headers[n]
    for h, c in headers.items():
        if any(a and a in h for a in av):
            continue
        for n in ns:
            if n and n in h:
                return c
    return None


def cell(row, col):
    if not col or col - 1 >= len(row):
        return ""
    return clean(row[col - 1])


def load_xlsx_table(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["PERIMETRO"] if "PERIMETRO" in wb.sheetnames else wb[wb.sheetnames[0]]
    hr = find_header(ws)
    max_col = safe_dim(ws.max_column, 80)
    headers = [norm_header(ws.cell(hr, c).value) for c in range(1, max_col + 1)]
    rows = list(ws.iter_rows(min_row=hr + 1, values_only=True))
    return headers, rows, path.name


def load_csv_table(path):
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=";,\t|")
    except Exception:
        dialect = csv.excel
        dialect.delimiter = ";"
    rows = list(csv.reader(text.splitlines(), dialect))
    header_idx = 0
    for i, row in enumerate(rows[:20]):
        hs = [norm_header(x) for x in row]
        if any(h in {"PV", "PDV", "PV ENI", "PUNTO EROGAZIONE"} or "PUNTO" in h for h in hs):
            header_idx = i
            break
    return [norm_header(x) for x in rows[header_idx]], rows[header_idx + 1:], path.name


def load_table(path):
    if not path:
        return [], [], ""
    if path.suffix.lower() == ".csv":
        return load_csv_table(path)
    return load_xlsx_table(path)


def load_lista():
    path = latest_file(INPUT_DIR / "lista", (".xlsx",))
    out = {}
    if not path:
        return out, ""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        pdv = norm_pdv(row[0] if row else None)
        if not pdv:
            continue
        out[pdv] = {
            "pdv": pdv,
            "agent": clean(row[9] if len(row) > 9 else ""),
            "region": clean(row[4] if len(row) > 4 else ""),
            "province": clean((row[5] if len(row) > 5 else "") or (row[2] if len(row) > 2 else "")),
            "city": clean((row[6] if len(row) > 6 else "") or (row[3] if len(row) > 3 else "")),
            "address": clean(row[7] if len(row) > 7 else ""),
            "source": "input/lista",
            "is_tp": True,
            "is_grab": False,
        }
    return out, path.name


def load_grab():
    path = latest_file(INPUT_DIR / "grabego")
    headers, rows, name = load_table(path)
    out = {}
    if not headers:
        return out, ""
    hmap = {h: i + 1 for i, h in enumerate(headers)}
    pdv_col = find_col(hmap, "PV", "PDV", "PV ENI", "PUNTO EROGAZIONE", "CODICE PV", "COD PV")
    agent_col = find_col(hmap, "AGENTE", "CONSULENTE", "SALES", "REFERENTE", "OWNER")
    region_col = find_col(hmap, "REGIONE")
    province_col = find_col(hmap, "PROVINCIA", "PROV")
    city_col = find_col(hmap, "CITTA", "COMUNE")
    address_col = find_col(hmap, "INDIRIZZO", "VIA")
    for row in rows:
        pdv = norm_pdv(cell(row, pdv_col))
        if not pdv:
            continue
        out[pdv] = {
            "pdv": pdv,
            "agent": cell(row, agent_col),
            "region": cell(row, region_col),
            "province": cell(row, province_col),
            "city": cell(row, city_col),
            "address": cell(row, address_col),
            "source": "input/grabego",
            "is_tp": False,
            "is_grab": True,
        }
    return out, name


def load_anag():
    path = latest_file(INPUT_DIR / "anagrafica", (".xlsx",))
    out = {}
    if not path:
        return out, ""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hr = find_header(ws)
    headers = headers_for_ws(ws, hr)
    cols = {
        "pdv": find_col(headers, "PV", "PDV", "PV ENI", "N PV", "N PV ENI", "PUNTO EROGAZIONE"),
        "region": find_col(headers, "REGIONE"),
        "province": find_col(headers, "PROVINCIA", "PROV"),
        "city": find_col(headers, "CITTA", "COMUNE"),
        "address": find_col(headers, "INDIRIZZO", "VIA"),
        "rzv": find_col(headers, "RZV"),
        "cr": find_col(headers, "CR"),
        "focal": find_col(headers, "FOCAL POINT ENI", "COORDINATORE SERVIZI", "COORD SERVIZI"),
        "lat": find_col(headers, "LATITUDINE", avoid=("VECCHIA", "NON VERIFICATA")),
        "lng": find_col(headers, "LONGITUDINE", avoid=("VECCHIA", "NON VERIFICATA")),
        "tpoint": find_col(headers, "TPOINT", "T POINT", "TELEPASS POINT"),
        "grab": find_col(headers, "GRABNGO", "GRAB NGO", "GRAB GO", "GRAB E GO"),
    }
    for row in ws.iter_rows(min_row=hr + 1, values_only=True):
        pdv = norm_pdv(cell(row, cols["pdv"]))
        if not pdv or pdv in out:
            continue
        lat = as_float(cell(row, cols["lat"]))
        lng = as_float(cell(row, cols["lng"]))
        gval = norm_header(cell(row, cols["grab"]))
        tval = norm_header(cell(row, cols["tpoint"]))
        out[pdv] = {
            "pdv": pdv,
            "agent": "",
            "region": cell(row, cols["region"]),
            "province": cell(row, cols["province"]),
            "city": cell(row, cols["city"]),
            "address": cell(row, cols["address"]),
            "rzv": cell(row, cols["rzv"]),
            "cr": cell(row, cols["cr"]),
            "focal": cell(row, cols["focal"]),
            "lat": lat,
            "lng": lng,
            "is_grab": gval in {"SI", "S", "YES", "TRUE", "1", "X"},
            "is_tp": tval not in {"", "NO", "N", "FALSE", "0"},
            "source": "anagrafica",
        }
    return out, path.name


def merge_point(base, extra):
    out = dict(base or {})
    for k, v in (extra or {}).items():
        if k in {"is_grab", "is_tp"}:
            out[k] = bool(out.get(k)) or bool(v)
        elif v not in (None, ""):
            out[k] = v
    return out


def build_planning_data():
    lista, lista_name = load_lista()
    grab, grab_name = load_grab()
    anag, anag_name = load_anag()
    active = {}
    for pdv, li in lista.items():
        p = merge_point(anag.get(pdv, {}), li)
        p["is_tp"] = True
        p["in_lista"] = True
        p["in_grabego"] = pdv in grab
        if pdv in grab:
            p = merge_point(p, grab[pdv])
            p["is_grab"] = True
            p["is_tp"] = True
            p["source"] = "input/lista + input/grabego"
        active[pdv] = p
    for pdv, gr in grab.items():
        if pdv in active:
            continue
        p = merge_point(anag.get(pdv, {}), gr)
        p["is_grab"] = True
        p["in_grabego"] = True
        p["in_lista"] = False
        active[pdv] = p

    def usable(p):
        return p.get("lat") is not None and p.get("lng") is not None

    points = [p for p in active.values() if usable(p)]
    no_coords = [p for p in active.values() if not usable(p)]
    catalog = list(anag.values())
    for p in catalog:
        p["in_lista"] = p.get("pdv") in lista
        p["in_grabego"] = p.get("pdv") in grab
    points.sort(key=lambda x: (x.get("agent") or "ZZZ", x.get("region") or "", x.get("province") or "", x.get("city") or "", x.get("pdv") or ""))
    catalog.sort(key=lambda x: x.get("pdv") or "")
    return {
        "source_name": f"lista: {lista_name or '-'} · grabego: {grab_name or '-'} · anagrafica: {anag_name or '-'}",
        "points": points,
        "catalog": catalog,
        "without_coordinates": no_coords,
        "summary": {
            "active_total": len(active),
            "active_with_coordinates": len(points),
            "active_without_coordinates": len(no_coords),
            "lista_count": len(lista),
            "grabego_count": len(grab),
            "catalog_count": len(catalog),
        },
    }


OVERLAY_CSS = """
<style>
.plan-manage-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(210px,1fr));gap:10px;margin-top:12px}.plan-mini-box{border:1px solid var(--line);border-radius:14px;padding:12px;background:#fff}.plan-mini-box h4{margin:0 0 8px 0;color:var(--blue)}.plan-mini-box input,.plan-mini-box select{width:100%;padding:8px;border:1px solid var(--line);border-radius:10px;margin:4px 0}.plan-danger{background:#fff5f5!important}.plan-added{background:#f0fff4!important}.plan-pill{display:inline-flex;gap:4px;align-items:center;border-radius:999px;padding:3px 8px;font-size:11px;font-weight:900;background:#edf2f7;color:#243447}.plan-pill.off{background:#ffe4e6;color:#9f1239}.plan-pill.on{background:#dcfce7;color:#166534}
</style>
"""

OVERLAY_JS = r'''
function planStorageJson(key, fallback){try{return JSON.parse(localStorage.getItem(key)||'');}catch(e){return fallback;}}
function planSaveJson(key, value){localStorage.setItem(key, JSON.stringify(value));}
function planCustomPoints(){return planStorageJson('planningCustomPoints', []);}
function planRemovedMap(){return planStorageJson('planningRemovedPdv', {});}
function planCatalog(){return (APP.planning_data&&APP.planning_data.catalog)||[];}
function planBasePoints(){return (APP.planning_data&&APP.planning_data.points)||[];}
function planAgents(){return [...new Set(planPoints().map(p=>p.agent).filter(Boolean))].sort((a,b)=>a.localeCompare(b,'it'));}
function planPoints(){const removed=planRemovedMap();const all=[...planBasePoints(),...planCustomPoints()];const by={};all.forEach(p=>{if(!p||!p.pdv||removed[p.pdv])return;by[p.pdv]=Object.assign(by[p.pdv]||{},p);});return Object.values(by);}
function planFindCatalog(pdv){const n=String(pdv||'').replace(/\D+/g,'').padStart(5,'0');return planCatalog().find(p=>p.pdv===n)||null;}
function planRefreshAgentOptions(){const sel=document.getElementById('planAgent');if(!sel)return;const cur=sel.value;sel.innerHTML='<option value="">Seleziona agente</option>'+planAgents().map(a=>`<option>${esc(a)}</option>`).join('');if(cur)sel.value=cur;}
function planRemovedCount(){return Object.keys(planRemovedMap()).length;}
function planCustomCount(){return planCustomPoints().length;}
function planManualSummary(){return `<span class="plan-pill on">Aggiunti: ${fmtNum(planCustomCount())}</span> <span class="plan-pill off">Esclusi: ${fmtNum(planRemovedCount())}</span>`;}
function renderPlanningAuto(){const w=document.getElementById('planningAutoWrap');if(!w)return;const agents=planAgents(),now=new Date(),month=`${now.getFullYear()}-${String(now.getMonth()+1).padStart(2,'0')}`,data=APP.planning_data||{},sum=data.summary||{};w.innerHTML=`<div class="card"><div class="section-title" style="margin-bottom:8px">Planning automatico</div><div class="small-muted">Fonte: ${esc(data.source_name||'')}<br>Usa solo PV presenti in <b>input/lista</b> + <b>input/grabego</b>. L'anagrafica completa serve solo come rubrica per coordinate/dati.</div><div class="metric-row sost plan-summary"><div class="metric-card"><h4>PV visitabili</h4><div class="metric-big">${fmtNum(sum.active_with_coordinates||planBasePoints().length)}</div><div class="metric-sub">Con coordinate</div></div><div class="metric-card ${sum.active_without_coordinates?'tone-warn':''}"><h4>Senza coordinate</h4><div class="metric-big">${fmtNum(sum.active_without_coordinates||0)}</div><div class="metric-sub">Da correggere in anagrafica</div></div><div class="metric-card"><h4>Modifiche locali</h4><div class="metric-big">${fmtNum(planCustomCount()+planRemovedCount())}</div><div class="metric-sub">${planManualSummary()}</div></div></div><div class="planning-form" style="margin-top:12px"><div><label>Agente</label><select id="planAgent"><option value="">Seleziona agente</option>${agents.map(a=>`<option>${esc(a)}</option>`).join('')}</select></div><div><label>Mese</label><input id="planMonth" type="month" value="${month}"></div><div><label>Punto di partenza</label><input id="planStart" placeholder="Città o PV di partenza"></div><div><label>Grab & Go</label><select id="planGrab"><option value="no">No</option><option value="yes">Sì, includili</option></select></div><div><label>Planning precedente</label><input id="planPrev" type="file" accept=".csv,.txt,.xls"></div></div><div class="plan-actions"><button class="btn" onclick="generatePlanning()">Crea planning</button><button class="btn light" onclick="downloadPlanCsv()">Scarica CSV</button><button class="btn light" onclick="downloadPlanXls()">Scarica Excel</button></div></div><div class="card"><div class="section-title" style="margin-bottom:6px">Gestione PV planning</div><div class="small-muted">Queste modifiche restano salvate nel browser di chi usa il sito. Non cambiano i file GitHub.</div><div class="plan-manage-grid"><div class="plan-mini-box plan-added"><h4>Aggiungi PV da anagrafica</h4><input id="planAddPdv" placeholder="Codice PV"><select id="planAddType"><option value="tp">Telepass Point</option><option value="grab">Grab & Go</option></select><button class="btn light small" onclick="addPlanningPdvFromCatalog()">Aggiungi</button><div id="planAddInfo" class="small-muted" style="margin-top:6px"></div></div><div class="plan-mini-box plan-danger"><h4>Escludi PV</h4><input id="planRemovePdv" placeholder="Codice PV da togliere"><button class="btn light small" onclick="removePlanningPdv()">Escludi</button><div class="small-muted" style="margin-top:6px">Utile per punti disattivati o da non visitare.</div></div><div class="plan-mini-box"><h4>Aggiunta manuale</h4><input id="manualPdv" placeholder="PV"><input id="manualCity" placeholder="Città"><input id="manualAddress" placeholder="Indirizzo"><input id="manualLat" placeholder="Latitudine"><input id="manualLng" placeholder="Longitudine"><select id="manualType"><option value="tp">Telepass Point</option><option value="grab">Grab & Go</option></select><button class="btn light small" onclick="addPlanningPdvManual()">Aggiungi manuale</button></div><div class="plan-mini-box"><h4>Reset modifiche</h4><div class="small-muted" style="margin-bottom:8px">${planManualSummary()}</div><button class="btn ghost small" onclick="resetPlanningLocalChanges()">Cancella modifiche locali</button></div></div></div><div id="planningResult"></div>`;document.getElementById('planPrev')?.addEventListener('change',e=>loadPrevPlanning(e.target.files[0]));document.getElementById('planAddPdv')?.addEventListener('input',previewPlanningPdv);}
function previewPlanningPdv(){const box=document.getElementById('planAddInfo');if(!box)return;const p=planFindCatalog(document.getElementById('planAddPdv')?.value);box.innerHTML=p?`Trovato: <b>${esc(p.pdv)}</b> · ${esc(p.city||'')} · ${esc(p.address||'')} · coord: ${p.lat&&p.lng?'ok':'mancanti'}`:'Non trovato in anagrafica';}
function addCustomPoint(p){if(!p||!p.pdv){alert('PV non valido');return;}if(p.lat==null||p.lng==null){alert('Questo PV non ha coordinate: non può entrare nel planning automatico');return;}const custom=planCustomPoints().filter(x=>x.pdv!==p.pdv);custom.push(p);planSaveJson('planningCustomPoints', custom);const removed=planRemovedMap();delete removed[p.pdv];planSaveJson('planningRemovedPdv', removed);renderPlanningAuto();}
function addPlanningPdvFromCatalog(){const pv=document.getElementById('planAddPdv')?.value||'',type=document.getElementById('planAddType')?.value||'tp',agent=document.getElementById('planAgent')?.value||'';const p=planFindCatalog(pv);if(!p){alert('PV non trovato in anagrafica. Usa aggiunta manuale.');return;}addCustomPoint({...p,agent:agent||p.agent||'',is_grab:type==='grab',is_tp:type!=='grab',source:'aggiunto manualmente da anagrafica'});}
function addPlanningPdvManual(){const agent=document.getElementById('planAgent')?.value||'';const pv=String(document.getElementById('manualPdv')?.value||'').replace(/\D+/g,'').padStart(5,'0');const lat=Number(String(document.getElementById('manualLat')?.value||'').replace(',','.')),lng=Number(String(document.getElementById('manualLng')?.value||'').replace(',','.'));if(!pv||!Number.isFinite(lat)||!Number.isFinite(lng)){alert('Servono almeno PV, latitudine e longitudine');return;}const type=document.getElementById('manualType')?.value||'tp';addCustomPoint({pdv,agent,city:document.getElementById('manualCity')?.value||'',address:document.getElementById('manualAddress')?.value||'',region:'',province:'',rzv:'',cr:'',focal:'',lat,lng,is_grab:type==='grab',is_tp:type!=='grab',source:'aggiunto manualmente'});}
function removePlanningPdv(){const pv=String(document.getElementById('planRemovePdv')?.value||'').replace(/\D+/g,'').padStart(5,'0');if(!pv){alert('Inserisci il codice PV');return;}const removed=planRemovedMap();removed[pv]=true;planSaveJson('planningRemovedPdv', removed);planSaveJson('planningCustomPoints', planCustomPoints().filter(p=>p.pdv!==pv));PLAN.items=(PLAN.items||[]).filter(p=>p.pdv!==pv);renderPlanningAuto();if(PLAN.items.length)renderPlanningTable(0,0);}
function resetPlanningLocalChanges(){if(!confirm('Vuoi cancellare aggiunte/esclusioni salvate su questo browser?'))return;localStorage.removeItem('planningCustomPoints');localStorage.removeItem('planningRemovedPdv');renderPlanningAuto();}
function generatePlanning(){const agent=document.getElementById('planAgent')?.value||'',month=document.getElementById('planMonth')?.value||'',includeGrab=(document.getElementById('planGrab')?.value==='yes'),startText=document.getElementById('planStart')?.value||'';if(!agent||!month){alert('Scegli agente e mese');return;}let pts=planPoints().filter(p=>p.agent===agent&&(includeGrab||!p.is_grab)&&p.lat!=null&&p.lng!=null);const noCoords=((APP.planning_data&&APP.planning_data.without_coordinates)||[]).filter(p=>p.agent===agent).length;if(!pts.length){alert('Nessun PV con coordinate per questo agente');return;}const days=planWorkdays(month);const start=planStartPoint(pts,startText);const ordered=planOrder(pts,start,month);PLAN.items=planAssign(ordered,days,start);PLAN.source=`${agent}_${month}`;renderPlanningTable(noCoords,days.length);}
'''


def replace_planning_data(html, data):
    payload = "APP.planning_data = " + json.dumps(data, ensure_ascii=False) + ";\nconst DATA"
    new, n = re.subn(r"APP\.planning_data\s*=\s*\{.*?\};\s*\nconst DATA", payload, html, count=1, flags=re.S)
    return new if n else html


def inject_overlay(html):
    if "planningCustomPoints" in html:
        return html
    if "</head>" in html:
        html = html.replace("</head>", OVERLAY_CSS + "\n</head>", 1)
    marker = "renderGarePdv();\nrenderGrabGo();\nrenderPlanningAuto();"
    if marker in html:
        html = html.replace(marker, OVERLAY_JS + "\n" + marker, 1)
    elif "</script>" in html:
        html = html.replace("</script>", OVERLAY_JS + "\n</script>", 1)
    return html


def main():
    data = build_planning_data()
    for name in ["index.html", "Telepass_ENI_sito_v6.html"]:
        path = DOCS_DIR / name
        if not path.exists():
            continue
        html = path.read_text(encoding="utf-8")
        html = replace_planning_data(html, data)
        html = inject_overlay(html)
        path.write_text(html, encoding="utf-8")
    print("Planning scope patch completata:", json.dumps(data.get("summary", {}), ensure_ascii=False))


if __name__ == "__main__":
    main()
